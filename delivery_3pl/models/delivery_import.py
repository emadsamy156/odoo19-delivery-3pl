import base64
import io
import json
import logging

from odoo import models, fields, api
from odoo.exceptions import ValidationError, UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("openpyxl not installed. Excel import will not work. Install with: pip install openpyxl")
    openpyxl = None

DEFAULT_COLUMN_MAPPING = {
    "row_num": ["م", "#", "Row", "row"],
    "rider_account_id": ["معرف الحساب", "account_id", "rider_account_id", "Account ID", "Rider ID", "rider_id"],
    "rider_name": ["اسم الحساب", "rider_name", "Rider Name", "Driver Name", "Name", "name"],
    "rider_user": ["مستخدم الحساب", "Account User", "User", "الاسم الحقيقي"],
    "vehicle_type_company": ["نوع المركبة من الشركة", "Vehicle Type Company", "vehicle_type"],
    "vehicle_type_contract": ["نوع المركبة حسب العقد", "Vehicle Type Contract"],
    "plate_number": ["رقم اللوحة", "Plate Number", "plate", "Plate"],
    "platform_target": ["تاجت كيتا", "تارجت كيتا", "Platform Target", "Keeta Target", "KT Target"],
    "company_target_val": ["التارجت الشركة", "تارجت الشركة", "Company Target", "CT Target"],
    "accepted_tasks": ["المهام المقبولة", "Accepted Tasks", "Accepted", "accepted"],
    "delivered_tasks": ["المهام التي تم تسليمها", "Delivered Tasks", "Delivered", "delivered"],
    "large_orders_completed": ["مهام الطلبات الكبيرة المكتملة", "Large Orders Completed", "Large Orders"],
    "cancelled_tasks": ["المهام المُلغاة", "المهام الملغاة", "Cancelled Tasks", "Cancelled", "cancelled"],
    "rejected_tasks": ["المهام المرفوضة", "Rejected Tasks", "Rejected", "rejected"],
    "driver_rejected": ["المهام المرفوضة (السائق)", "Driver Rejected", "driver_rejected"],
    "auto_rejected": ["المهام المرفوضة تلقائيًا (تلقائياً)", "المهام المرفوضة تلقائيًا", "Auto Rejected", "auto_rejected"],
    "online_hours": ["وقت اتصال السائقين عبر تطبيق السائق.", "وقت اتصال السائقين", "Online Hours", "online_hours", "Connection Time"],
    "ontime_rate": ["نسبة الطلبات التي تم تسليمها في الوقت المحدد (D)", "نسبة التسليم في الوقت", "On-time Rate", "ontime_rate", "OTD"],
    "large_ontime_rate": ["معدل توصيل الطلبات الكبيرة في الوقت المُحدَّد", "Large Order On-time", "large_ontime_rate"],
    "avg_delivery_duration": ["متوسط مدة التوصيل لكل طلب مكتمل", "متوسط مدة التوصيل", "Avg Duration", "avg_duration"],
    "over_55min_rate": ["نسبة الطلبات المُسلمة (أكثر من 55 دقيقة).", "نسبة أكثر من 55 دقيقة", "Over 55min Rate"],
    "late_tasks": ["مهام الطلبات المتأخرة", "Late Tasks", "late_tasks"],
    "very_late_tasks": ["مهام الطلبات المتأخرة جدًا", "Very Late Tasks", "very_late_tasks"],
    "fuel": ["بنزين", "Fuel", "fuel", "Gas"],
    "order_id": ["order_id", "order id", "رقم الطلب", "Order ID", "Order No"],
    "rider_phone": ["phone", "rider_phone", "هاتف", "رقم الجوال", "جوال", "Phone", "Mobile"],
    "city_name": ["city", "city_name", "المدينة", "City", "City Name"],
    "order_date": ["date", "order_date", "التاريخ", "تاريخ الطلب", "Order Date"],
    "distance": ["distance", "distance_km", "المسافة", "Distance", "Distance (km)", "KM"],
    "platform_amount": ["amount", "platform_amount", "المبلغ", "مبلغ المنصة", "Amount", "Platform Amount", "Total"],
}


class DeliveryImportSession(models.Model):
    _name = 'delivery.import.session'
    _description = 'Delivery Import Session'
    _inherit = ['mail.thread']
    _order = 'import_date desc'

    company_id = fields.Many2one('delivery.company', string='Company', required=True, tracking=True, ondelete='cascade')
    branch_id = fields.Many2one('delivery.company.branch', string='Branch', tracking=True,
                                 domain="[('company_id', '=', company_id)]")
    contract_id = fields.Many2one('delivery.contract', string='Contract', tracking=True,
                                  domain="[('company_id', '=', company_id), ('status', '=', 'active'), '|', ('branch_id', '=', branch_id), ('branch_id', '=', False)]")
    file_name = fields.Char(string='File Name', tracking=True)
    file_data = fields.Binary(string='Excel File', attachment=True)
    import_date = fields.Date(string='Import Date', required=True, default=fields.Date.today, tracking=True)
    period_start = fields.Date(string='Period Start', required=True)
    period_end = fields.Date(string='Period End', required=True)
    sheet_name = fields.Char(string='Sheet Name (اسم الورقة)',
                             help='اسم الورقة المراد قراءتها. اتركه فارغاً لقراءة الورقة الأولى.\n'
                                  'Sheet name to read. Leave empty to read the first sheet.')
    status = fields.Selection([
        ('pending', 'Pending'),
        ('validated', 'Validated'),
        ('failed', 'Failed'),
        ('imported', 'Imported'),
    ], string='Status', default='pending', required=True, tracking=True)
    total_rows = fields.Integer(string='Total Rows', default=0)
    valid_rows = fields.Integer(string='Valid Rows', default=0)
    error_rows = fields.Integer(string='Error Rows', default=0)
    total_amount = fields.Float(string='Total Amount', digits=(12, 2), default=0.0)
    imported_by = fields.Many2one('res.users', string='Imported By', default=lambda self: self.env.user)
    notes = fields.Text(string='Notes')
    error_log = fields.Text(string='Error Log')

    row_ids = fields.One2many('delivery.import.row', 'session_id', string='Import Rows')
    row_count = fields.Integer(compute='_compute_row_count')

    def _compute_row_count(self):
        for rec in self:
            rec.row_count = len(rec.row_ids)

    @api.onchange('company_id')
    def _onchange_company_id(self):
        if self.company_id:
            self.branch_id = False
            active_contract = self.env['delivery.contract'].search([
                ('company_id', '=', self.company_id.id),
                ('status', '=', 'active'),
            ], limit=1)
            self.contract_id = active_contract.id if active_contract else False

    @api.onchange('branch_id')
    def _onchange_branch_id(self):
        if self.branch_id:
            active_contract = self.env['delivery.contract'].search([
                ('company_id', '=', self.company_id.id),
                ('branch_id', '=', self.branch_id.id),
                ('status', '=', 'active'),
            ], limit=1)
            if active_contract:
                self.contract_id = active_contract.id

    def _get_column_mapping(self):
        self.ensure_one()
        if self.contract_id and self.contract_id.column_mapping:
            try:
                custom_mapping = json.loads(self.contract_id.column_mapping)
                if isinstance(custom_mapping, dict):
                    return custom_mapping
            except (json.JSONDecodeError, TypeError):
                _logger.warning("Invalid column_mapping JSON in contract %s, using defaults", self.contract_id.id)
        return DEFAULT_COLUMN_MAPPING

    def _detect_columns(self, headers_combined):
        mapping = self._get_column_mapping()
        detected = {}
        for col_idx, header_val in enumerate(headers_combined):
            if not header_val:
                continue
            header_str = str(header_val).strip()
            header_lower = header_str.lower()
            for field_name, aliases in mapping.items():
                if field_name in detected:
                    continue
                alias_list = aliases if isinstance(aliases, list) else [aliases]
                for alias in alias_list:
                    alias_str = str(alias).strip()
                    if header_lower == alias_str.lower() or header_str == alias_str:
                        detected[field_name] = col_idx
                        break
        return detected

    def _build_merged_headers(self, ws, max_col):
        row1 = []
        row2 = []
        for col_idx in range(max_col):
            v1 = ws.cell(row=1, column=col_idx + 1).value
            v2 = ws.cell(row=2, column=col_idx + 1).value
            row1.append(str(v1).strip() if v1 is not None else None)
            row2.append(str(v2).strip() if v2 is not None else None)

        has_sub_headers = any(
            row2[i] is not None and row1[i] is None
            for i in range(max_col)
        )

        if has_sub_headers:
            merged = []
            for i in range(max_col):
                if row2[i] is not None:
                    merged.append(row2[i])
                elif row1[i] is not None:
                    merged.append(row1[i])
                else:
                    merged.append(None)
            data_start_row = 3
        else:
            merged = row1
            data_start_row = 2

        return merged, data_start_row

    def action_parse_file(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError('الرجاء رفع ملف Excel أولاً. / Please upload an Excel file first.')
        if not openpyxl:
            raise UserError('مكتبة openpyxl غير مثبتة. الرجاء تثبيتها: pip install openpyxl')

        self.row_ids.unlink()

        try:
            file_content = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            self.write({
                'status': 'failed',
                'error_log': f'فشل في قراءة الملف / Failed to read file: {str(e)}',
            })
            return

        if self.sheet_name and self.sheet_name.strip():
            target_sheet = self.sheet_name.strip()
            if target_sheet in wb.sheetnames:
                ws = wb[target_sheet]
            else:
                available = ', '.join(wb.sheetnames)
                wb.close()
                self.write({
                    'status': 'failed',
                    'error_log': f'الورقة "{target_sheet}" غير موجودة.\n'
                                 f'الأوراق المتاحة: {available}\n\n'
                                 f'Sheet "{target_sheet}" not found.\n'
                                 f'Available sheets: {available}',
                })
                return
        else:
            ws = wb[wb.sheetnames[0]]

        sheet_used = ws.title
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0

        if max_row < 2 or max_col < 1:
            wb.close()
            self.write({
                'status': 'failed',
                'error_log': f'الورقة "{sheet_used}" فارغة أو لا تحتوي على بيانات كافية.\n'
                             f'Sheet "{sheet_used}" is empty or has insufficient data.',
            })
            return

        headers_combined, data_start_row = self._build_merged_headers(ws, max_col)

        col_map = self._detect_columns(headers_combined)

        if not col_map:
            header_names = [h for h in headers_combined if h]
            wb.close()
            self.write({
                'status': 'failed',
                'error_log': (
                    f'الورقة المستخدمة: {sheet_used}\n'
                    f'لم يتم التعرف على أي عمود. / No columns were detected.\n'
                    f'أعمدة الملف / File headers: {", ".join(header_names)}\n\n'
                    f'الرجاء ضبط خريطة الأعمدة (Column Mapping) في العقد.\n'
                    f'Please configure Column Mapping in the contract.'
                ),
            })
            return

        rider_model = self.env['delivery.rider']
        city_model = self.env['delivery.city']
        import_row_model = self.env['delivery.import.row']

        total = 0
        valid = 0
        errors = 0
        error_lines = []
        total_amount = 0.0
        rows_to_create = []

        for row_idx in range(data_start_row, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                row_data.append(ws.cell(row=row_idx, column=col_idx).value)

            if all(v is None or str(v).strip() == '' for v in row_data):
                continue

            total += 1
            row_vals = {
                'session_id': self.id,
                'row_number': total,
                'status': 'pending',
            }

            def get_val(field):
                idx = col_map.get(field)
                if idx is not None and idx < len(row_data):
                    return row_data[idx]
                return None

            def get_str(field):
                v = get_val(field)
                return str(v).strip() if v is not None and str(v).strip() not in ('', 'None') else False

            def get_float(field):
                v = get_val(field)
                if v is None:
                    return 0.0
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return 0.0

            row_vals['order_id'] = get_str('order_id') or get_str('rider_account_id') or False
            row_vals['rider_phone'] = get_str('rider_phone') or False
            row_vals['rider_name'] = get_str('rider_name') or False
            row_vals['rider_user'] = get_str('rider_user') or False
            row_vals['city_name'] = get_str('city_name') or False
            row_vals['vehicle_type_company'] = get_str('vehicle_type_company') or False
            row_vals['vehicle_type_contract'] = get_str('vehicle_type_contract') or False
            row_vals['plate_number'] = get_str('plate_number') or False

            row_vals['platform_target'] = get_float('platform_target')
            row_vals['company_target_val'] = get_float('company_target_val')
            row_vals['accepted_tasks'] = int(get_float('accepted_tasks'))
            row_vals['delivered_tasks'] = int(get_float('delivered_tasks'))
            row_vals['large_orders_completed'] = int(get_float('large_orders_completed'))
            row_vals['cancelled_tasks'] = int(get_float('cancelled_tasks'))
            row_vals['rejected_tasks'] = int(get_float('rejected_tasks'))
            row_vals['driver_rejected'] = int(get_float('driver_rejected'))
            row_vals['auto_rejected'] = int(get_float('auto_rejected'))
            row_vals['online_hours'] = get_float('online_hours')
            row_vals['ontime_rate'] = get_float('ontime_rate')
            row_vals['large_ontime_rate'] = get_float('large_ontime_rate')
            row_vals['avg_delivery_duration'] = get_float('avg_delivery_duration')
            row_vals['over_55min_rate'] = get_float('over_55min_rate')
            row_vals['late_tasks'] = int(get_float('late_tasks'))
            row_vals['very_late_tasks'] = int(get_float('very_late_tasks'))
            row_vals['fuel'] = get_float('fuel')

            order_date_val = get_val('order_date')
            if order_date_val:
                try:
                    if hasattr(order_date_val, 'date'):
                        row_vals['order_date'] = order_date_val.date()
                    elif hasattr(order_date_val, 'strftime'):
                        row_vals['order_date'] = order_date_val
                    else:
                        from datetime import datetime
                        parsed = datetime.strptime(str(order_date_val).strip(), '%Y-%m-%d')
                        row_vals['order_date'] = parsed.date()
                except (ValueError, TypeError):
                    row_vals['order_date'] = False

            row_vals['distance'] = get_float('distance')

            amount_val = get_float('platform_amount')
            row_vals['platform_amount'] = amount_val
            total_amount += amount_val

            row_errors = []
            rider_account_id = get_str('rider_account_id') or False
            matched_rider = False

            if rider_account_id:
                rider = rider_model.search([
                    ('platform_account_id', '=', rider_account_id),
                    ('primary_company_id', '=', self.company_id.id),
                ], limit=1)
                if not rider:
                    rider = rider_model.search([
                        ('platform_account_id', '=', rider_account_id),
                    ], limit=1)
                if rider:
                    matched_rider = rider.id

            if not matched_rider and row_vals.get('rider_phone'):
                rider = rider_model.search([
                    ('phone', '=', row_vals['rider_phone']),
                    ('primary_company_id', '=', self.company_id.id),
                ], limit=1)
                if not rider:
                    rider = rider_model.search([
                        ('phone', '=', row_vals['rider_phone']),
                    ], limit=1)
                if rider:
                    matched_rider = rider.id

            search_names = []
            if row_vals.get('rider_user'):
                search_names.append(row_vals['rider_user'])
            if row_vals.get('rider_name') and row_vals['rider_name'] != '--':
                search_names.append(row_vals['rider_name'])

            for sname in search_names:
                if matched_rider:
                    break
                rider = rider_model.search([
                    '|',
                    ('name', '=', sname),
                    ('name_ar', '=', sname),
                ], limit=1)
                if rider:
                    matched_rider = rider.id

            row_vals['rider_id'] = matched_rider
            if not matched_rider:
                identifier = rider_account_id or row_vals.get('rider_phone') or row_vals.get('rider_user') or row_vals.get('rider_name') or f'Row {total}'
                row_errors.append(f'مندوب غير معروف / Rider not found: {identifier}')

            matched_city = False
            if row_vals.get('city_name'):
                city = city_model.search([
                    '|',
                    ('name', 'ilike', row_vals['city_name']),
                    ('name_ar', 'ilike', row_vals['city_name']),
                ], limit=1)
                if city:
                    matched_city = city.id
            row_vals['city_id'] = matched_city

            if row_errors:
                row_vals['status'] = 'error'
                row_vals['error_message'] = '; '.join(row_errors)
                errors += 1
                error_lines.append(f"صف {total}: {'; '.join(row_errors)}")
            else:
                row_vals['status'] = 'valid'
                valid += 1

            rows_to_create.append(row_vals)

        wb.close()

        if rows_to_create:
            import_row_model.create(rows_to_create)

        detected_cols = [f"{k} → Col {v+1}: {headers_combined[v]}" for k, v in sorted(col_map.items(), key=lambda x: x[1])]

        update_vals = {
            'total_rows': total,
            'valid_rows': valid,
            'error_rows': errors,
            'total_amount': total_amount,
        }

        log_parts = [
            f'الورقة المستخدمة / Sheet used: {sheet_used}',
            f'بداية البيانات / Data starts at row: {data_start_row}',
            f'',
            f'الأعمدة المكتشفة / Detected columns ({len(col_map)}):',
        ] + detected_cols

        if error_lines:
            log_parts += ['', f'أخطاء / Errors ({errors}):'] + error_lines[:100]
            if len(error_lines) > 100:
                log_parts.append(f'... و {len(error_lines) - 100} أخطاء أخرى')
        else:
            log_parts += ['', 'لا توجد أخطاء / No errors found.']

        update_vals['error_log'] = '\n'.join(log_parts)
        self.write(update_vals)

        self._auto_create_performance()

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'delivery.import.session',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def _auto_create_performance(self):
        import logging
        _logger = logging.getLogger(__name__)

        perf_model = self.env['delivery.daily.performance']
        created = 0
        updated = 0
        perf_errors = []

        valid_rows = self.row_ids.filtered(lambda r: r.rider_id)
        _logger.info('3PL Import: _auto_create_performance called for session %s with %d rows (rider matched)', self.id, len(valid_rows))

        if not valid_rows:
            _logger.warning('3PL Import: No rows with matched riders found for session %s', self.id)
            return

        for row in valid_rows:
            try:
                perf_date = row.order_date or self.period_start
                if not perf_date:
                    perf_date = fields.Date.context_today(self)

                existing = perf_model.search([
                    ('rider_id', '=', row.rider_id.id),
                    ('date', '=', perf_date),
                    ('branch_id', '=', self.branch_id.id),
                ], limit=1)

                perf_vals = {
                    'branch_id': self.branch_id.id,
                    'rider_id': row.rider_id.id,
                    'date': perf_date,
                    'import_session_id': self.id,
                    'platform_account_id': row.order_id or False,
                    'account_name': row.rider_user or row.rider_name or False,
                    'vehicle_type_company': self._normalize_vehicle_type(row.vehicle_type_company),
                    'vehicle_type_contract': self._normalize_vehicle_type(row.vehicle_type_contract),
                    'license_plate': row.plate_number or False,
                    'platform_target': int(row.platform_target or 0),
                    'accepted_orders': int(row.accepted_tasks or 0),
                    'delivered_orders': int(row.delivered_tasks or 0),
                    'large_orders_completed': int(row.large_orders_completed or 0),
                    'cancelled_orders': int(row.cancelled_tasks or 0),
                    'rejected_orders': int(row.rejected_tasks or 0),
                    'total_online_hours': float(row.online_hours or 0.0),
                }

                if existing:
                    existing.write(perf_vals)
                    updated += 1
                else:
                    perf_model.create(perf_vals)
                    created += 1

                rider_updates = {}
                if row.order_id and not row.rider_id.platform_account_id:
                    rider_updates['platform_account_id'] = row.order_id
                if row.vehicle_type_contract:
                    vt = self._normalize_vehicle_type(row.vehicle_type_contract)
                    if vt and not row.rider_id.vehicle_type:
                        rider_updates['vehicle_type'] = vt
                if row.plate_number and not row.rider_id.license_plate:
                    rider_updates['license_plate'] = row.plate_number
                if rider_updates:
                    row.rider_id.write(rider_updates)

            except Exception as e:
                _logger.error('3PL Import: Error creating performance for row %s rider %s: %s',
                              row.row_number, row.rider_id.name_ar or row.rider_id.name, str(e))
                perf_errors.append(f'Row {row.row_number}: {str(e)}')

        perf_summary = f'\n\nأداء يومي / Daily Performance: إنشاء {created} / تحديث {updated}'
        if perf_errors:
            perf_summary += f'\nأخطاء أداء ({len(perf_errors)}):\n' + '\n'.join(perf_errors[:20])
        current_log = self.error_log or ''
        self.write({'error_log': current_log + perf_summary})

    def action_validate(self):
        self.ensure_one()
        if self.status != 'pending':
            raise ValidationError('Only pending imports can be validated.')
        if not self.row_ids:
            raise ValidationError(
                'لا توجد صفوف للتحقق. الرجاء معالجة الملف أولاً بالضغط على "معالجة الملف".\n'
                'No rows to validate. Please process the file first by clicking "Process File".'
            )
        error_count = len(self.row_ids.filtered(lambda r: r.status == 'error'))
        valid_count = len(self.row_ids.filtered(lambda r: r.status == 'valid'))
        pending_count = len(self.row_ids.filtered(lambda r: r.status == 'pending'))
        self.write({
            'status': 'validated',
            'valid_rows': valid_count,
            'error_rows': error_count,
        })

    def _normalize_vehicle_type(self, raw_val):
        if not raw_val:
            return False
        val = str(raw_val).strip().lower()
        if val in ('car', 'private car', 'سيارة'):
            return 'car'
        if val in ('bike', 'motorcycle', 'دراجة', 'موتور'):
            return 'motorcycle'
        return False

    def action_confirm(self):
        self.ensure_one()
        if self.status != 'validated':
            raise ValidationError('Only validated imports can be confirmed.')
        self._auto_create_performance()
        self.write({'status': 'imported'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'delivery.import.session',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_mark_failed(self):
        self.ensure_one()
        self.write({'status': 'failed'})

    def action_reset_to_pending(self):
        self.ensure_one()
        self.row_ids.unlink()
        self.write({
            'status': 'pending',
            'total_rows': 0,
            'valid_rows': 0,
            'error_rows': 0,
            'total_amount': 0.0,
            'error_log': False,
        })


class DeliveryImportRow(models.Model):
    _name = 'delivery.import.row'
    _description = 'Import Row'
    _order = 'row_number'

    session_id = fields.Many2one('delivery.import.session', string='Import Session', required=True, ondelete='cascade')
    row_number = fields.Integer(string='Row #', required=True)
    order_id = fields.Char(string='Order / Account ID')
    rider_phone = fields.Char(string='Rider Phone')
    rider_name = fields.Char(string='Rider Name (اسم الحساب)')
    rider_user = fields.Char(string='Account User (مستخدم الحساب)')
    rider_id = fields.Many2one('delivery.rider', string='Matched Rider')
    city_name = fields.Char(string='City Name')
    city_id = fields.Many2one('delivery.city', string='Matched City')
    order_date = fields.Date(string='Order Date')

    vehicle_type_company = fields.Char(string='Vehicle Type (Company)')
    vehicle_type_contract = fields.Char(string='Vehicle Type (Contract)')
    plate_number = fields.Char(string='Plate Number (رقم اللوحة)')

    platform_target = fields.Float(string='Platform Target (تارجت المنصة)', digits=(8, 2))
    company_target_val = fields.Float(string='Company Target (تارجت الشركة)', digits=(8, 2))

    accepted_tasks = fields.Integer(string='Accepted Tasks (المهام المقبولة)')
    delivered_tasks = fields.Integer(string='Delivered Tasks (المهام المسلّمة)')
    large_orders_completed = fields.Integer(string='Large Orders Completed')
    cancelled_tasks = fields.Integer(string='Cancelled Tasks (الملغاة)')
    rejected_tasks = fields.Integer(string='Rejected Tasks (المرفوضة)')
    driver_rejected = fields.Integer(string='Driver Rejected (رفض السائق)')
    auto_rejected = fields.Integer(string='Auto Rejected (رفض تلقائي)')

    online_hours = fields.Float(string='Online Hours (ساعات الاتصال)', digits=(8, 2))
    ontime_rate = fields.Float(string='On-time Rate (نسبة التسليم بالوقت)', digits=(5, 4))
    large_ontime_rate = fields.Float(string='Large Order On-time Rate', digits=(5, 4))
    avg_delivery_duration = fields.Float(string='Avg Delivery Duration (min)', digits=(8, 2))
    over_55min_rate = fields.Float(string='Over 55min Rate', digits=(5, 4))
    late_tasks = fields.Integer(string='Late Tasks (متأخرة)')
    very_late_tasks = fields.Integer(string='Very Late Tasks (متأخرة جداً)')

    fuel = fields.Float(string='Fuel (بنزين)', digits=(8, 2))

    distance = fields.Float(string='Distance (km)', digits=(8, 2))
    platform_amount = fields.Float(string='Platform Amount', digits=(12, 2))
    calculated_amount = fields.Float(string='Calculated Amount', digits=(12, 2))
    variance = fields.Float(string='Variance', digits=(12, 2))

    status = fields.Selection([
        ('pending', 'Pending'),
        ('valid', 'Valid'),
        ('error', 'Error'),
    ], string='Status', default='pending')
    error_message = fields.Char(string='Error Message')
